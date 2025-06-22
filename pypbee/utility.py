# -*- coding: utf-8 -*-
"""
Created on Thu Nov  7 15:43:05 2019

@author: Angshuman Deb
"""

from .multivariate_nataf import multivariate_nataf
import os
import platform
import numpy as np
import importlib.util
from scipy.stats import norm
import subprocess
import pickle
from openpyxl import load_workbook
import xlrd
import string
import eqsig
from datetime import datetime
from benedict import benedict
import matplotlib.pyplot as plt
import matplotlib.colors as colors
from matplotlib import cm
from mpl_toolkits.axes_grid1.inset_locator import inset_axes


class Utility:
    ####################################################################################################################
    # Namespace class with static methods
    ####################################################################################################################

    @staticmethod
    def str_replace(str_from, str_to, file_to_modify, path_of_file):
        if str_from is None:
            return
        file_to_modify = Utility.get_path(path_of_file, file_to_modify)
        with open(file_to_modify, 'r') as fid:
            lines = fid.readlines()
        line_inds = [ind for ind in range(len(lines)) if str_from in lines[ind]]
        for ind in line_inds:
            lines[ind] = lines[ind].replace(str_from, str_to)
        with open(file_to_modify, 'w') as fid:
            fid.writelines(lines)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def rgb(c):
        return colors.to_rgb(c)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def block_diag_3d(*args):
        depth = args[0].shape[2]
        shapes = np.array([a.shape[:-1] for a in args])
        out = np.zeros(tuple(np.sum(shapes, axis=0)) + (depth,))
        r, c = 0, 0
        for i, (rr, cc) in enumerate(shapes):
            out[r:r + rr, c:c + cc, :] = args[i]
            r += rr
            c += cc
        return out

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def is_list_of_same_elems(lst):
        return all(ele == lst[0] for ele in lst)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def plot_raw_ground_acc_time_history(gm_database_dir_path, **kwargs):
        rec_num = kwargs.get('rec_num', None)
        gm_file_names = kwargs.get('gm_file_names', None)
        scale_fac = kwargs.get('scale_fac', 1.0)
        if rec_num is not None and gm_file_names is None:
            # Extract rec_data
            rec_data_file_path = Utility.get_path(os.path.dirname(__file__), 'Rec_Data', 'rec_data.pickle')
            rec_data = Utility.pickle_load_dict(rec_data_file_path)
            file_name_1 = rec_data['file_name_1']
            file_name_2 = rec_data['file_name_2']
            file_name_vert = rec_data['file_name_vert']
            file_ind = rec_num - 1
            gm_file_names = [file_name_1[file_ind], file_name_2[file_ind], file_name_vert[file_ind]]
            gm_file_paths = [Utility.get_path(gm_database_dir_path, gm_file_name) for gm_file_name in gm_file_names]
        elif gm_file_names is not None and rec_num is None:
            gm_file_paths = [Utility.get_path(gm_database_dir_path, gm_file_name) for gm_file_name in gm_file_names]
        else:
            raise ValueError("either rec_num or gm_file_names should be not NoneType")
        grid_alpha = kwargs.get('grid_alpha', 0.5)
        minor_grid_alpha = kwargs.get('minor_grid_alpha', 0.0)
        lc = kwargs.get('lc', 'black')
        max_marker = kwargs.get('max_marker', 'o')
        mec = kwargs.get('mec', 'red')
        ms = kwargs.get('ms', 5)
        lw = kwargs.get('lw', 0.8)
        fs = kwargs.get('fs', 8)
        txc = kwargs.get('txc', 'red')
        max_text_precision = kwargs.get('max_text_precision', 5.3)
        fig, axs = plt.subplots(len(gm_file_paths), 1)
        for itr in range(len(gm_file_paths)):
            gm, npts, dt = Utility.read_nga_peer_eq_at2_file(gm_file_paths[itr])
            if gm is not None:
                t = np.linspace(0.0, dt * npts, int(npts))
                axs[itr].plot(t, gm * scale_fac, '-', color=lc, linewidth=lw)
                axs[itr].minorticks_on()
                axs[itr].grid(True, which="major", alpha=grid_alpha)
                axs[itr].grid(True, which="minor", alpha=minor_grid_alpha)
                axs[itr].set_title(gm_file_names[itr] + f', scale: {scale_fac:.3}', fontsize=fs)
                axs[itr].tick_params(axis='x', labelsize=fs)
                axs[itr].tick_params(axis='y', labelsize=fs)
                Utility.mark_max_in_time_hist(axs[itr], gm * scale_fac, t, max_marker,
                                              markeredgecolor=mec,
                                              markersize=ms,
                                              precision=max_text_precision,
                                              units='g',
                                              textcolor=txc,
                                              fontsize=fs,
                                              ha='left',
                                              va='center')
            else:
                axs[itr].set_title('File NOT found', fontsize=fs)
        return fig, axs

    @staticmethod
    def fetch_selected_ground_motion_records(gm_database_dir_path, gm_file_names, scale_fac, target_dir_path,
                                             ai_end):
        time_hist_list = list()
        npts_list = list()
        dt_list = list()
        t_start_list = list()
        t_end_list = list()
        for (i, gm_file) in zip(list(range(1, len(gm_file_names) + 1)), gm_file_names):
            time_hist_temp, npts_temp, dt_temp = Utility.read_nga_peer_eq_at2_file(
                Utility.get_path(gm_database_dir_path, gm_file))
            time_hist_list.append(time_hist_temp)
            npts_list.append(npts_temp)
            dt_list.append(dt_temp)
            if npts_temp is not None:
                t_sig = eqsig.im.calc_sig_dur(eqsig.AccSignal(time_hist_temp * 9.81, dt_temp),
                                              start=0.001, end=ai_end,
                                              se=True)
            else:
                t_sig = [None, None]
            t_start_list.append(t_sig[0])
            t_end_list.append(t_sig[1])

        if npts_list[-1] is None:
            npts_list[-1] = npts_list[0]
            dt_list[-1] = dt_list[0]
            time_hist_list[-1] = time_hist_list[0] * 0.
            t_start_list[-1] = t_start_list[0]
            t_end_list[-1] = t_end_list[0]

        t_start = min(t_start_list)
        t_end = max(t_end_list)
        for i in range(len(gm_file_names)):
            npts_start = round(t_start / dt_list[i])
            npts_end = round(t_end / dt_list[i])
            time_hist = time_hist_list[i][int(npts_start):int(npts_end)]
            save_arr = np.vstack([npts_end - npts_start, dt_list[i], time_hist.reshape(len(time_hist), 1) * scale_fac])
            Utility.save_array_as_text_file(Utility.get_path(target_dir_path, f'gm_{i + 1}.txt'), save_arr)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def save_array_as_text_file(filepath, array, **kwargs):
        np.savetxt(Utility.get_path(filepath), array, **kwargs)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def read_nga_peer_eq_at2_file(file_path):
        file_path = file_path.replace('.at2', '.AT2')

        if not os.path.isfile(file_path):
            return None, None, None

        with open(file_path, 'r') as fid:
            skipheader = 0
            go_on = True
            while go_on:
                curr_line = fid.readline()
                words = curr_line.split()
                if all(Utility.is_numeric_string(word) for word in words):
                    go_on = False
                    num_cols = len(words)
                else:
                    skipheader += 1

        with open(file_path, 'r') as fid:
            lines = fid.readlines()

        line_index = -1
        go_on = True
        while go_on:
            curr_line = lines[line_index].strip().split()
            if len(curr_line) == num_cols:
                go_on = False
            else:
                line_index -= 1
        skipfooter = (-line_index) - 1

        time_hist = np.loadtxt(file_path, skiprows=skipheader, max_rows=len(lines) - skipheader - skipfooter).flatten()
        if line_index != -1:
            last_row = np.array(lines[line_index + 1].split()).astype(float)
            time_hist = np.hstack((time_hist, last_row))

        temp = [float(word) for word in lines[skipheader - 1].replace(',', ' ').replace(';', ' ').split() if
                Utility.is_numeric_string(word)]
        npts = int(max(temp))
        dt = float(min(temp))

        return time_hist, npts, dt

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_tag_of_token(name, token):
        name_parts = name.split('/')
        name_parts_with_token = [name_part for name_part in name_parts if token in name_part.split('_')]
        tags = list()
        for name_part_with_token in name_parts_with_token:
            parts_of_name_part_with_token = name_part_with_token.split('_')
            if '.' in parts_of_name_part_with_token[-1]:
                parts_of_name_part_with_token[-1] = parts_of_name_part_with_token[-1].split('.')[0]
            inds = [i for i in range(len(parts_of_name_part_with_token)) if parts_of_name_part_with_token[i] == token]
            tags.extend([parts_of_name_part_with_token[ind + 1] for ind in inds if
                         len(parts_of_name_part_with_token) != ind + 1
                         and Utility.is_numeric_string(parts_of_name_part_with_token[ind + 1])])
        if len(tags) > 0:
            return tags[0]
        else:
            return ''

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def exist_line_in_file(fid, line_to_test):
        line_to_test = line_to_test.strip('\n')
        pos = fid.tell()
        fid.seek(0)

        lines = fid.readlines()

        try:
            lines.index(f'{line_to_test}\n')
        except ValueError:
            try:
                lines.index(f'{line_to_test};\n')
            except ValueError:
                line_found = False
                fid.seek(pos)
                return line_found

        line_found = True
        fid.seek(pos)
        return line_found

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def pickle_dump_dict(file_path, dictionary):
        if type(dictionary) == benedict:
            dictionary = dict(dictionary)
        with open(file_path, 'wb') as f:
            pickle.dump(dictionary, f)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def pickle_load_dict(file_path):
        try:
            with open(file_path, 'rb') as f:
                dictionary = pickle.load(f)
            return dictionary
        except FileNotFoundError:
            raise FileNotFoundError

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def in_nd(a, b):

        a = list(a)
        b = list(b)

        a = [list(item) for item in a]
        b = [list(item) for item in b]

        to_return = [True if item in b else False for item in a]

        return np.array(to_return)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def combine_contents_of_files(dir_path, which_files, out_file_name, **kwargs):

        # which_files = ['all', 'startswith <tag>', 'contains <tag>', 'file_1', 'file_2', ...]

        file_ext = kwargs.get('file_ext', '.txt')
        delete_original = kwargs.get('delete_original', False)
        out_file_open_mode = kwargs.get('out_file_open_mode', 'w')

        files_to_cat = list()

        stuff_in_dir_path = os.listdir(dir_path)

        for file_names in which_files:
            if type(file_names) == str and file_names == 'all':
                files_to_cat.extend([stuff for stuff in stuff_in_dir_path if stuff.endswith(file_ext)])

            elif type(file_names) == str and file_names.startswith('startswith '):
                tag = file_names.split()[1]
                files_to_cat.extend(
                    [stuff for stuff in stuff_in_dir_path if stuff.startswith(tag) and stuff.endswith(file_ext)])

            elif type(file_names) == str and file_names.startswith('contains '):
                tags = file_names.split()[1:]
                files_to_cat.extend(
                    [stuff for stuff in stuff_in_dir_path if all(tag in stuff for tag in tags)
                     and stuff.endswith(file_ext)])

            elif type(file_names) == str:
                files_to_cat.append(file_names)

        files_to_cat = list(np.unique(files_to_cat))
        lines = Utility.get_lines_in_files(dir_path, files_to_cat)

        if len(files_to_cat) > 0:
            if '\\' in out_file_name or '/' in out_file_name:
                out_file_path = out_file_name
            else:
                out_file_path = Utility.get_path(dir_path, out_file_name)
            with open(out_file_path, out_file_open_mode) as fid:
                fid.writelines(lines)

        if len(files_to_cat) == 1 and files_to_cat[0] == out_file_name:
            delete_original = False

        if delete_original:
            [os.remove(Utility.get_path(dir_path, file)) for file in files_to_cat]

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_lines_in_files(dir_path, files_to_cat):
        lines = list()
        for file in files_to_cat:
            with open(Utility.get_path(dir_path, file), 'r') as fid:
                lines.extend(fid.readlines())
                lines[-1] = lines[-1].strip('\n')
                lines[-1] += '\n'

        return lines

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def make_array_2d(array):
        if array.ndim == 1:
            return array.reshape(1, array.shape[0])
        else:
            return array

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def make_array_3d(array):
        if array.ndim == 1:
            return array.reshape(1, array.shape[0], 1)
        elif array.ndim == 2:
            return array.reshape(array.shape[0], array.shape[1], 1)
        else:
            return array

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def write_tcl_input_file(fid, param_values, param_names):
        for itr in range(len(param_values)):
            if type(param_values[itr]) != str:
                fid.write(f'set {param_names[itr]} {param_values[itr]};\n')
            else:
                fid.write(f'set {param_names[itr]} "{param_values[itr]}";\n')

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_multi_level_iterator_cases(*args):
        # Input args should be lists or 1d array_like of unique numeric elements!
        vec = list()
        for arg in args:
            if type(arg) != list and type(arg) != np.ndarray:
                arg = [arg]
            arg = np.array(arg).reshape(len(arg))
            vec.append(arg)

        itr_range_size = [len(i) for i in vec]
        num_cases = np.prod(itr_range_size)
        itr_range_size_padded = [1, *itr_range_size, 1]
        cases = np.zeros((num_cases, len(itr_range_size)))
        curr_level = -1

        while curr_level >= -len(vec):
            temp = np.repeat(np.array(vec[curr_level]), np.prod(itr_range_size_padded[curr_level:]))
            temp = temp.reshape(len(temp), 1)
            temp = np.tile(temp, (np.prod(itr_range_size_padded[:curr_level - 1]), 1))
            cases[:, curr_level] = temp[:, 0]
            curr_level -= 1

        return cases

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_num(list_of_ints):
        return int(''.join([str(i) for i in list_of_ints]))

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_standard_beta_dist_params(mean, std, min_val, max_val):
        mu = mean / (max_val - min_val) - min_val / (max_val - min_val)
        sig = 1 / (max_val - min_val) * std
        a = -mu * np.divide(np.square(sig) + np.square(mu) - mu, np.square(sig))
        b = (np.square(sig) + np.square(mu) - mu) * np.divide(mu - 1, np.square(sig))
        return a, b

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_standard_normal_clips(myclip_a, myclip_b, my_mean, my_std):
        return (myclip_a - my_mean) / my_std, (myclip_b - my_mean) / my_std

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_my_normal_clips(a, b, my_mean, my_std):
        return a * my_std + my_mean, b * my_std + my_mean

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_lognormal_dist_params(mean, std, **kwargs):
        loc = kwargs.get('loc', np.abs(0.0 * mean))
        mean -= loc
        sig = np.sqrt(np.log(1 + np.square(std / np.abs(mean))))
        mu = np.log(mean) - 0.5 * np.square(sig)
        return sig, loc, np.exp(mu)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_all_dist_params(prob_dist):
        if prob_dist.dist.name == 'mixture':
            return prob_dist.args
        temp = prob_dist.dist.__dict__['_parse_args'](*prob_dist.args, **prob_dist.kwds)
        return [*temp[0], *temp[1:len(temp)]]

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def setup_rng(rng_seed):
        if (type(rng_seed) == str) and rng_seed.startswith('unique'):
            set_rng_seed = True
            temp = rng_seed.split(' ')
            if len(temp) > 1:
                seed_multiplier = int(float(temp[1]))
            else:
                seed_multiplier = 1
        else:
            set_rng_seed = False
            seed_multiplier = 1

        return set_rng_seed, seed_multiplier

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def import_attr_from_module(module_path, module_name, attr):
        module_path = Utility.get_path(module_path, module_name + '.py')
        spec = importlib.util.spec_from_file_location(module_name, module_path)
        temp = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(temp)
        return getattr(temp, attr)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_path(*args):
        args = [arg.replace('\\', '/') if isinstance(arg, str) else None for arg in args]
        args = list(filter(None, args))
        return '/'.join(args)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def is_numeric_string(n):
        try:
            # Type-casting the string to `float`.
            # If string is not a valid `float`, 
            # it'll raise `ValueError` exception
            float(n)
        except ValueError:
            return False
        return True

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def read_numpy_array_from_txt_file(file_path, **kwargs):

        # skiprows = any int, 0 (default), or 'find'

        skiprows = kwargs.get('skiprows', 0)
        comments = kwargs.get('comments', '#')

        if type(skiprows) == str and skiprows == 'find':
            with open(file_path, 'r') as fid:
                skiprows = 0
                go_on = True
                while go_on:
                    curr_line = fid.readline()
                    words = curr_line.split()
                    if all(Utility.is_numeric_string(word) for word in words) and len(words) > 0:
                        go_on = False
                    else:
                        skiprows += 1

        return Utility.make_array_2d(np.loadtxt(file_path, skiprows=skiprows, comments=comments))

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_indexed_matrix(m, ind):
        n = len(ind)
        inds = np.arange(m.shape[0])
        m_ind = np.in1d(inds, ind)
        m_ind = np.outer(m_ind, m_ind)
        return m[m_ind].reshape((n, n))

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def generate_random_dist_param_vals(prob_dist_list, corr_matrix, estimation_sample_size, dist_params_sample_size,
                                        **kwargs):

        rng_seed = kwargs.get('rng_seed', None)

        nd = len(prob_dist_list)
        out_list = list()
        m = multivariate_nataf(prob_dist_list, corr_matrix)
        m_rvs = m.rvs(size=(dist_params_sample_size, estimation_sample_size),
                      random_state=rng_seed,
                      method='mcs').reshape((dist_params_sample_size, estimation_sample_size, nd))
        m_rvs_mean = np.mean(m_rvs, axis=1)[:, np.newaxis, :]
        m_rvs_std = np.std(m_rvs, axis=1, ddof=1)
        m_rvs_temp = m_rvs - m_rvs_mean
        corr_matrix_realizations = (1 / (estimation_sample_size - 1) * (
            np.einsum('ijk,ijl->ikl', m_rvs_temp, m_rvs_temp))) / np.einsum('ij,ik->ijk', m_rvs_std, m_rvs_std)
        corr_matrix = corr_matrix[np.newaxis, :, :].repeat(dist_params_sample_size, axis=0)
        corr_matrix_realizations[corr_matrix == 0] = 0
        corr_matrix_realizations[corr_matrix == 1] = 1
        m_rvs_mean = m_rvs_mean.squeeze(axis=1)

        for itr in range(nd):
            temp = Utility.get_all_dist_params(prob_dist_list[itr])

            if prob_dist_list[itr].dist.name == 'beta':
                beta_a, beta_b = Utility.get_standard_beta_dist_params(m_rvs_mean[:, itr], m_rvs_std[:, itr], temp[2],
                                                                       temp[3] + temp[2])
                out_list.append(np.column_stack(
                    (beta_a, beta_b, [temp[2]] * dist_params_sample_size, [temp[3]] * dist_params_sample_size)))

            if prob_dist_list[itr].dist.name == 'lognorm':
                lognorm_zeta, lognorm_loc, lognorm_med = Utility.get_lognormal_dist_params(
                    m_rvs_mean[:, itr], m_rvs_std[:, itr], loc=[temp[1]] * dist_params_sample_size)
                out_list.append(np.column_stack((lognorm_zeta, lognorm_loc, lognorm_med)))

            if prob_dist_list[itr].dist.name == 'norm':
                out_list.append(np.column_stack((m_rvs_mean[:, itr], m_rvs_std[:, itr])))

            if prob_dist_list[itr].dist.name == 'truncnorm':
                clip_a, clip_b = Utility.get_standard_normal_clips(*Utility.get_my_normal_clips(*temp),
                                                                   m_rvs_mean[:, itr], m_rvs_std[:, itr])
                out_list.append(np.column_stack((clip_a, clip_b, m_rvs_mean[:, itr], m_rvs_std[:, itr])))

        return out_list, corr_matrix_realizations.transpose((1, 2, 0))

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def fit_dist_to_data(fit_dist, data):
        dist_name = fit_dist.name
        if dist_name == 'lognorm':
            return fit_dist.fit(data, floc=0)
        else:
            return fit_dist.fit(data)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def is_pos_def(x):
        return np.all(np.linalg.eigvals(x) > 0)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_cell_range(sheet, start_col, start_row, end_col, end_row):
        return [[cell.value for cell in sheet.row_slice(row, start_colx=start_col, end_colx=end_col + 1)] for row in
                range(start_row, end_row + 1)]

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def col2index(col):
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num - 1

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def xlsread(file_path, sheet_name, data_range):
        data_range = data_range.split(':')
        if '.xlsx' in file_path:
            wb = load_workbook(file_path, read_only=True)
            ws = wb[sheet_name]
            # Read the cell values into a list of lists
            data_rows = list()
            for row in ws[data_range[0]:data_range[1]]:
                data_cols = list()
                for cell in row:
                    data_cols.append(cell.value)
                data_rows.append(data_cols)
        elif '.xls' in file_path:
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_name(sheet_name)
            start_col = ''
            for c in data_range[0]:
                if not Utility.is_numeric_string(c):
                    start_col += c
            start_col = Utility.col2index(start_col)
            start_row = ''
            for c in data_range[0]:
                if Utility.is_numeric_string(c):
                    start_row += c
            start_row = int(start_row) - 1
            end_col = ''
            for c in data_range[1]:
                if not Utility.is_numeric_string(c):
                    end_col += c
            end_col = Utility.col2index(end_col)
            end_row = ''
            for c in data_range[1]:
                if Utility.is_numeric_string(c):
                    end_row += c
            end_row = int(end_row) - 1
            data_rows = Utility.get_cell_range(ws, start_col, start_row, end_col, end_row)
        else:
            return

        return np.array(data_rows)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_exact_mixture_params(mean_req, sig_req, weights):
        vector_size = mean_req.shape[1]
        mean_req_exact = np.sum(
            mean_req * Utility.make_array_2d(weights).T, axis=0)
        var_req_exact = np.sum(
            Utility.make_array_2d(weights).T *
            (np.square(sig_req) + np.square(mean_req - mean_req_exact)),
            axis=0)
        sig_req_exact = np.sqrt(var_req_exact)
        return mean_req_exact.reshape(1, vector_size), sig_req_exact.reshape(1, vector_size)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def calc_oscillator_resp(freq,
                             fourier_amp,
                             osc_damping,
                             osc_freq,
                             max_freq_ratio=5.,
                             peak_resp_only=False,
                             osc_type='psa'):
        """Compute the time series response of an oscillator.
    
        Parameters
        ----------
        freq : array_like
            frequency of the Fourier acceleration spectrum [Hz]
        fourier_amp : array_like
            Fourier acceleration spectrum [g-sec]
        osc_damping : float
            damping of the oscillator [decimal]
        osc_freq : float
            frequency of the oscillator [Hz]
        max_freq_ratio : float, default=5
            minimum required ratio between the oscillator frequency and
            then maximum frequency of the time series. It is recommended that this
            value be 5.
        peak_resp_only : bool, default=False
            If only the peak response is returned.
        osc_type : str, default='psa'
            type of response. Options are:
                'sd': spectral displacement
                'sv': spectral velocity
                'sa': spectral acceleration
                'psv': psuedo-spectral velocity
                'psa': psuedo-spectral acceleration
        Returns
        -------
        response : :class:`numpy.ndarray` or float
            time series response of the oscillator
        
        Copyright 2016-18 Albert Kottke - pyrotd
        
        """
        ang_freq = 2 * np.pi * freq
        osc_ang_freq = 2 * np.pi * osc_freq

        # Single-degree of freedom transfer function
        h = (1 / (ang_freq ** 2. - osc_ang_freq ** 2 -
                  2.j * osc_damping * osc_ang_freq * ang_freq))
        if osc_type == 'sd':
            pass
        elif osc_type == 'sv':
            h *= 1.j * ang_freq
        elif osc_type == 'sa':
            h *= -(ang_freq ** 2)
        elif osc_type == 'psa':
            h *= -(osc_ang_freq ** 2)
        elif osc_type == 'psv':
            h *= -osc_ang_freq
        else:
            raise RuntimeError

        # Adjust the maximum frequency considered. The maximum frequency is 5
        # times the oscillator frequency. This provides that at the oscillator
        # frequency there are at least tenth samples per wavelength.
        n = len(fourier_amp)
        m = max(n, int(max_freq_ratio * osc_freq / freq[1]))
        scale = float(m) / float(n)

        # Scale factor is applied to correct the amplitude of the motion for the
        # change in number of points
        resp = scale * np.fft.irfft(fourier_amp * h, 2 * (m - 1))

        if peak_resp_only:
            resp = np.abs(resp).max()

        return resp

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_response_spectra(gm_data, dt, **kwargs):
        periods = np.array(kwargs.get('periods', np.logspace(np.log10(0.05), np.log10(5), 50, endpoint=True))).flatten()
        xi = kwargs.get('xi', 0.05)

        fourier_amp = np.fft.rfft(gm_data)
        freq = np.linspace(0, 1. / (2 * dt), num=fourier_amp.size)
        spec_accels = [Utility.calc_oscillator_resp(
            freq, fourier_amp, xi, 1 / period, max_freq_ratio=5,
            peak_resp_only=True, osc_type='psa'
        ) for period in periods]

        return np.column_stack((periods, spec_accels))

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def replace_pattern_in_file(file_path, pattern_to_find, pattern_to_replace):
        with open(file_path, 'r') as fid:
            lines = fid.readlines()
        lines = [line.replace(pattern_to_find, pattern_to_replace) for line in lines]
        with open(file_path, 'w') as fid:
            fid.writelines(lines)
        return

    @staticmethod
    def replace_in_files(dir_path, pattern_to_find, pattern_to_replace, ext):
        files = os.listdir(dir_path)
        for file in files:
            if file.split('.')[-1] == ext.split('.')[-1]:
                Utility.replace_pattern_in_file(Utility.get_path(dir_path, file), pattern_to_find, pattern_to_replace)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def mark_max_in_time_hist(ax, time_hist, t, marker, **kwargs):
        markeredgecolor = kwargs.get('markeredgecolor', 'r')
        markersize = kwargs.get('markersize', 10)
        precision = kwargs.get('precision', 5.3)
        units = kwargs.get('units', '')
        textcolor = kwargs.get('textcolor', 'k')
        fontsize = kwargs.get('fontsize', 12)
        ha = kwargs.get('ha', 'left')
        va = kwargs.get('va', 'center')

        ind_max = np.argmax(np.abs(time_hist))
        ax.plot(t[ind_max], time_hist[ind_max], marker, markerfacecolor="None", markeredgecolor=markeredgecolor,
                markersize=markersize)
        ax.text(t[ind_max], time_hist[ind_max], f'{np.abs(time_hist[ind_max]):{precision}} {units}', color=textcolor,
                fontsize=fontsize, horizontalalignment=ha, verticalalignment=va)
        return ax

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def add_common_labels_to_sub_plots(fig, xlabel, ylabel):
        # add a big axis, hide frame
        fig.add_subplot(111, frameon=False)
        ax = fig.gca()
        # hide tick and tick label of the big axis
        ax.tick_params(labelcolor='none', top=False, bottom=False, left=False, right=False)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        return

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_xyz_within_lims(x, y, z, x_lim, y_lim, z_lim):
        mask = np.all([x >= x_lim[0], x <= x_lim[-1], y >= y_lim[0], y <= y_lim[-1], z >= z_lim[0], z <= z_lim[-1]],
                      axis=0)
        return x[mask], y[mask], z[mask]

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_xy_within_lims(x, y, x_lim, y_lim):
        mask = np.all([x >= x_lim[0], x <= x_lim[-1], y >= y_lim[0], y <= y_lim[-1]], axis=0)
        return x[mask], y[mask]

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def merge_sorted_1d_arrays(big_array, small_array):
        if np.isscalar(small_array):
            small_array = np.array([small_array])
        big_array = np.array(big_array)
        small_array = np.array(small_array)
        for i_s in range(len(small_array)):
            if not np.any(big_array == small_array[i_s]):
                big_array = np.hstack(
                    (big_array[big_array < small_array[i_s]],
                     small_array[i_s],
                     big_array[big_array > small_array[i_s]]))

        return big_array

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def ecdf(data):
        x = np.sort(data)
        n = x.size
        y = np.arange(1, n + 1) / n
        return x, y

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_total_correlation_estimate(mixture_list, corr_matrix, rng_seed=None):
        # mixture_list : list of marginal mixtures, each with n_s components
        # corr_matrix : (n_rv, n_rv, <n_s>) n_s correlation matrices

        mixture_size = mixture_list[0].mixture_size
        weights = mixture_list[0].weights

        for itr in range(1, len(mixture_list)):
            if np.any(mixture_list[itr].weights != weights):
                return

        if len(corr_matrix.shape) == 2:
            corr_matrix = corr_matrix[:, :, np.newaxis].repeat(mixture_size, axis=2)

        np.random.seed(rng_seed)
        mixture_inds = np.random.choice(mixture_size, size=500, p=weights)
        rvs_temp = np.array([]).reshape(0, len(mixture_list))
        for itr in range(len(mixture_inds)):
            ind = mixture_inds[itr]
            prob_dist_list = [item.prob_dists[ind] if item.each == 'arbitrary'
                              else item.each(*np.array(item.args)[ind, :]) for item in mixture_list]
            m = multivariate_nataf(prob_dist_list, corr_matrix[:, :, ind])
            if rng_seed is not None:
                rng_seed_send = Utility.get_num([rng_seed, itr])
            else:
                rng_seed_send = rng_seed
            rvs_temp = np.vstack((rvs_temp,
                                  m.rvs(size=1, random_state=rng_seed_send, method='mcs')
                                  ))

        return np.corrcoef(rvs_temp, rowvar=False)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def plot_2d_sampling_steps(prob_dist_list, corr_matrix, sample_size, sampling_method, **kwargs):
        rng_seed = kwargs.get('rng_seed', None)
        figkwargs = kwargs.get('figkwargs', dict())
        grid_alpha = kwargs.get('grid_alpha', 0.5)
        minor_grid_alpha = kwargs.get('minor_grid_alpha', 0.0)
        k = kwargs.get('k', 5)
        n_cont = kwargs.get('n_cont', 100)
        ms = kwargs.get('ms', 10)
        lc = kwargs.get('lc', 'black')
        lw = kwargs.get('lw', 1)
        mc = kwargs.get('mc', 'red')
        n_bins = kwargs.get('n_bins', 10)
        hc = kwargs.get('hc', 'gray')
        same_lim = kwargs.get('same_lim', False)
        dist_func = kwargs.get('dist_func', 'pdf')
        fig_ax = kwargs.get('fig_ax', None)
        scilimits = kwargs.get('scilimits', (-4, 4))

        num = figkwargs.get('num', None)
        figkwargs['num'] = num

        m = multivariate_nataf(prob_dist_list, corr_matrix)
        x, independent_samples, dependent_samples, uniform_dependent_samples, transformed_samples = m.sampler(
            sample_size, rng_seed, sampling_method, return_all=True
        )

        if sampling_method == 'lhs':
            grid, ind = m.get_2d_lhs_grid(sample_size, n_cont)
            x_grid, u_grid, z_grid, x_star_grid, y_grid = grid
        else:
            ind, x_grid, u_grid, z_grid, x_star_grid, y_grid = (None,) * 6

        export_mat_dict = dict()
        if fig_ax is None or np.array(fig_ax, dtype='object').any() is None:
            fig_1 = plt.figure(**figkwargs)
            ax_1 = fig_1.gca()
        else:
            fig_1 = fig_ax[0][0]
            ax_1 = fig_ax[1][0]
        ax_1.minorticks_on()
        ax_1.grid(True, which="major", alpha=grid_alpha)
        ax_1.grid(True, which="minor", alpha=minor_grid_alpha)
        ax_1.plot(x[:, 0], x[:, 1], '.', markersize=ms, color=mc)
        ax_1.ticklabel_format(axis='both', scilimits=scilimits)
        if sampling_method == 'lhs':
            ax_1.grid(False, which="both")
            ax_1 = Utility.plot_grid(ax_1, x_grid[:, :, 0], x_grid[:, :, 1], ind, lw, lc, grid_alpha)
            export_mat_dict['ind'] = ind
            export_mat_dict['x_grid'] = x_grid
        ax_1.set_xlim((0, 1))
        ax_1.set_ylim((0, 1))
        export_mat_dict['x'] = x

        figkwargs['num'] = fig_1.number + 1
        if fig_ax is None or np.array(fig_ax, dtype='object').any() is None:
            fig_ax_2 = None
        else:
            fig_ax_2 = [fig_ax[0][1], fig_ax[1][1]]
        fig_2, ax_2, temp_dict = Utility.scatter_hist_2d(independent_samples,
                                                         multivariate_nataf([norm(), norm()], np.eye(2),
                                                                            corr_matrix_z=np.eye(2)),
                                                         figkwargs=figkwargs,
                                                         grid_alpha=grid_alpha,
                                                         minor_grid_alpha=minor_grid_alpha,
                                                         k=k, n_cont=n_cont,
                                                         n_bins=n_bins, ms=ms, mc=mc,
                                                         lw=lw, lc=lc, hc=hc,
                                                         same_lim=same_lim, dist_func=dist_func, fig_ax=fig_ax_2)
        ax_2[0].ticklabel_format(axis='both', scilimits=scilimits)
        if sampling_method == 'lhs':
            ax_2[0].grid(False, which="both")
            ax_2[0] = Utility.plot_grid(ax_2[0], u_grid[:, :, 0], u_grid[:, :, 1], ind, lw, lc, grid_alpha)
            export_mat_dict['u_grid'] = u_grid
        export_mat_dict['u'] = independent_samples
        export_mat_dict[f'u_{dist_func}_plot'] = np.dstack((temp_dict[f'{dist_func}_plot_1'],
                                                            temp_dict[f'{dist_func}_plot_2']))
        export_mat_dict['u_joint_pdf_surf'] = temp_dict[f'joint_pdf_surf']

        figkwargs['num'] = fig_2.number + 1
        if fig_ax is None or np.array(fig_ax, dtype='object').any() is None:
            fig_ax_3 = None
        else:
            fig_ax_3 = [fig_ax[0][2], fig_ax[1][2]]
        fig_3, ax_3, temp_dict = Utility.scatter_hist_2d(dependent_samples,
                                                         multivariate_nataf([norm(), norm()], m.corr_matrix_z,
                                                                            corr_matrix_z=m.corr_matrix_z),
                                                         figkwargs=figkwargs,
                                                         grid_alpha=grid_alpha,
                                                         minor_grid_alpha=minor_grid_alpha,
                                                         k=k, n_cont=n_cont,
                                                         n_bins=n_bins, ms=ms, mc=mc,
                                                         lw=lw, lc=lc, hc=hc,
                                                         same_lim=same_lim, dist_func=dist_func, fig_ax=fig_ax_3)
        ax_3[0].ticklabel_format(axis='both', scilimits=scilimits)
        if sampling_method == 'lhs':
            ax_3[0].grid(False, which="both")
            ax_3[0] = Utility.plot_grid(ax_3[0], z_grid[:, :, 0], z_grid[:, :, 1], ind, lw, lc, grid_alpha)
            export_mat_dict['z_grid'] = z_grid
        export_mat_dict['z'] = dependent_samples
        export_mat_dict[f'z_{dist_func}_plot'] = np.dstack((temp_dict[f'{dist_func}_plot_1'],
                                                            temp_dict[f'{dist_func}_plot_2']))
        export_mat_dict['z_joint_pdf_surf'] = temp_dict[f'joint_pdf_surf']
        export_mat_dict['z_corr_matrix'] = m.corr_matrix_z

        figkwargs['num'] = fig_3.number + 1
        if fig_ax is None or np.array(fig_ax, dtype='object').any() is None:
            fig_4 = plt.figure(**figkwargs)
            ax_4 = fig_4.gca()
        else:
            fig_4 = fig_ax[0][3]
            ax_4 = fig_ax[1][3]
        ax_4.minorticks_on()
        ax_4.grid(True, which="major", alpha=grid_alpha)
        ax_4.grid(True, which="minor", alpha=minor_grid_alpha)
        ax_4.ticklabel_format(axis='both', scilimits=scilimits)
        ax_4.plot(uniform_dependent_samples[:, 0], uniform_dependent_samples[:, 1], '.', markersize=ms, color=mc)
        if sampling_method == 'lhs':
            ax_4.grid(False, which="both")
            ax_4 = Utility.plot_grid(ax_4, x_star_grid[:, :, 0], x_star_grid[:, :, 1], ind, lw, lc, grid_alpha)
            export_mat_dict['x*_grid'] = x_star_grid
        ax_4.set_xlim((0, 1))
        ax_4.set_ylim((0, 1))
        export_mat_dict['x*'] = uniform_dependent_samples

        figkwargs['num'] = fig_4.number + 1
        if fig_ax is None or np.array(fig_ax, dtype='object').any() is None:
            fig_ax_5 = None
        else:
            fig_ax_5 = [fig_ax[0][4], fig_ax[1][4]]
        fig_5, ax_5, temp_dict = Utility.scatter_hist_2d(transformed_samples, m,
                                                         figkwargs=figkwargs,
                                                         grid_alpha=grid_alpha,
                                                         minor_grid_alpha=minor_grid_alpha,
                                                         k=k, n_cont=n_cont,
                                                         n_bins=n_bins, ms=ms, mc=mc,
                                                         lw=lw, lc=lc, hc=hc,
                                                         same_lim=same_lim, dist_func=dist_func, fig_ax=fig_ax_5)
        ax_5[0].ticklabel_format(axis='both', scilimits=scilimits)
        if sampling_method == 'lhs':
            ax_5[0].grid(False, which="both")
            ax_5[0] = Utility.plot_grid(ax_5[0], y_grid[:, :, 0], y_grid[:, :, 1], ind, lw, lc, grid_alpha)
            export_mat_dict['y_grid'] = y_grid
        export_mat_dict['y'] = transformed_samples
        export_mat_dict[f'y_{dist_func}_plot'] = np.dstack((temp_dict[f'{dist_func}_plot_1'],
                                                            temp_dict[f'{dist_func}_plot_2']))
        export_mat_dict['y_joint_pdf_surf'] = temp_dict[f'joint_pdf_surf']

        return [fig_1, fig_2, fig_3, fig_4, fig_5], [ax_1, ax_2, ax_3, ax_4, ax_5], export_mat_dict

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def plot_grid(ax, xx, yy, ind, lw, lc, grid_alpha):
        for itr in range(1, len(ind) - 1):
            ax.plot(xx[:, ind[itr]], yy[:, ind[itr]], '-', linewidth=lw, color=lc, alpha=grid_alpha, zorder=1)
            ax.plot(xx[ind[itr], :], yy[ind[itr], :], '-', linewidth=lw, color=lc, alpha=grid_alpha, zorder=1)
        return ax

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def scatter_hist_2d(data, mvnataf, **kwargs):
        figkwargs = kwargs.get('figkwargs', dict())
        grid_alpha = kwargs.get('grid_alpha', 0.5)
        minor_grid_alpha = kwargs.get('minor_grid_alpha', 0.0)
        k = kwargs.get('k', 5)
        n_cont = kwargs.get('n_cont', 100)
        n_bins = kwargs.get('n_bins', 10)
        ms = kwargs.get('ms', 10)
        lc = kwargs.get('lc', 'black')
        lw = kwargs.get('lw', 1)
        hc = kwargs.get('hc', 'gray')
        mc = kwargs.get('mc', 'red')
        same_lim = kwargs.get('same_lim', False)
        dist_func = kwargs.get('dist_func', 'pdf')
        fig_ax = kwargs.get('fig_ax', None)

        left, width = 0.1, 0.65
        bottom, height = 0.1, 0.65
        spacing = 0.005

        rect_scatter = [left, bottom, width, height]
        rect_histx = [left, bottom + height + spacing, width, 0.2]
        rect_histy = [left + width + spacing, bottom, 0.2, height]

        # start with a rectangular Figure
        if fig_ax is None or np.array(fig_ax, dtype='object').any() is None:
            fig = plt.figure(**figkwargs)
            axs = list()
            axs.append(fig.add_axes(rect_scatter))
            axs.append(fig.add_axes(rect_histx))
            axs.append(fig.add_axes(rect_histy))
        else:
            fig = fig_ax[0]
            axs = fig_ax[1]

        axs[0].minorticks_on()
        axs[0].grid(True, which="major", alpha=grid_alpha)
        axs[0].grid(True, which="minor", alpha=minor_grid_alpha)
        axs[0].tick_params(direction='in', top=True, right=True)
        axs[1].tick_params(direction='in', labelbottom=False)
        axs[2].tick_params(direction='in', labelleft=False)

        # the scatter plot:
        x = data[:, 0]
        y = data[:, 1]
        axs[0].plot(x, y, '.', color=mc, markersize=ms)
        if dist_func == 'cdf':
            cumulative = True
        else:
            cumulative = False

        axs[1].hist(x, bins=n_bins, density=True, color=hc, edgecolor='white', cumulative=cumulative)
        axs[2].hist(y, bins=n_bins, orientation='horizontal', density=True, color=hc, edgecolor='white',
                    cumulative=cumulative)

        export_mat_dict = dict()
        prob_dist_list = mvnataf.prob_dist_list
        for itr in range(len(prob_dist_list)):
            ax = axs[itr + 1]
            prob_dist = prob_dist_list[itr]
            mean = prob_dist.mean()
            std = prob_dist.std()
            x1 = mean - k * std
            x2 = mean + k * std
            x = np.linspace(x1, x2, n_cont)
            if dist_func == 'cdf':
                df_x = prob_dist.cdf(x)
            else:
                df_x = prob_dist.pdf(x)
            if itr == 0:
                ax.plot(x, df_x, color=lc, linewidth=lw)
            if itr == 1:
                ax.plot(df_x, x, color=lc, linewidth=lw)

            ax.axis('off')
            export_mat_dict[f'{dist_func}_plot_{itr + 1}'] = np.column_stack((x, df_x))
            export_mat_dict[f'dist_name_{itr + 1}'] = prob_dist.dist.name

        export_mat_dict[f'data'] = data

        mu_1 = prob_dist_list[0].mean()
        mu_2 = prob_dist_list[1].mean()
        sig_1 = prob_dist_list[0].std()
        sig_2 = prob_dist_list[1].std()
        xlim = (mu_1 - k * sig_1, mu_1 + k * sig_1)
        ylim = (mu_2 - k * sig_2, mu_2 + k * sig_2)
        x_bounds = prob_dist_list[0].support()
        y_bounds = prob_dist_list[1].support()
        xlim = (max(xlim[0], x_bounds[0]), min(xlim[1], x_bounds[1]))
        ylim = (max(ylim[0], y_bounds[0]), min(ylim[1], y_bounds[1]))

        x_surf = np.linspace(*xlim, n_cont)
        y_surf = np.linspace(*ylim, n_cont)
        x_surf = x_surf[1:]
        y_surf = y_surf[1:]
        surf_xx, surf_yy = np.meshgrid(x_surf, y_surf)
        zz = mvnataf.pdf(np.dstack((surf_xx, surf_yy)))
        axs[0].contour(surf_xx, surf_yy, zz)
        export_mat_dict[f'joint_pdf_surf'] = np.dstack([surf_xx, surf_yy, zz])

        if same_lim:
            xlim = (min(xlim[0], ylim[0]), max(xlim[1], ylim[1]))
            ylim = xlim
        axs = Utility.adjust_scatter_hist_2d_axes(axs, xlim, ylim)

        return fig, axs, export_mat_dict

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def adjust_scatter_hist_2d_axes(axs, xlim, ylim):
        axs[0].set_xlim(xlim)
        axs[0].set_ylim(ylim)
        axs[1].set_xlim(axs[0].get_xlim())
        axs[2].set_ylim(axs[0].get_ylim())
        return axs

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def plot_3d_bar(x, y, dz, **kwargs):
        figkwargs = kwargs.get('figkwargs', dict())
        cmap = kwargs.get('cmap', 'viridis')
        minor_grid = kwargs.get('minor_grid', False)
        fig_ax = kwargs.get('fig_ax', None)

        if fig_ax is None or np.array(fig_ax, dtype='object').any() is None:
            fig = plt.figure(**figkwargs)  # also accepts tight_layout=bool
            ax = fig.add_subplot(projection='3d')
        else:
            fig = fig_ax[0]
            ax = fig_ax[1]
        if minor_grid:
            ax.minorticks_on()
        else:
            ax.minorticks_off()

        temp = np.diff(x)
        dx = np.ones(len(x)) * min(temp[temp > 0])

        temp = np.diff(y)
        dy = np.ones(len(y)) * min(temp[temp > 0])

        fracs = dz / dz.max()
        normalize = colors.Normalize(fracs.min(), fracs.max())
        my_cmap = cm.get_cmap(cmap)
        color_values = my_cmap(normalize(fracs.tolist()))

        z = np.zeros(len(x))

        ax.bar3d(x, y, z, dx, dy, dz, color=color_values)
        ax_in = inset_axes(ax, width="50%", height="5%", loc='upper right')

        sm = cm.ScalarMappable(cmap=my_cmap)
        sm.set_array(dz)
        fig.colorbar(sm, cax=ax_in, orientation='horizontal')

        export_mat_dict = dict()
        export_mat_dict['x'] = x
        export_mat_dict['y'] = y
        export_mat_dict['dz'] = dz

        return fig, ax, export_mat_dict

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def scatter_hist_1d(data, prob_dist=None, **kwargs):
        figkwargs = kwargs.get('figkwargs', dict())
        grid_alpha = kwargs.get('grid_alpha', 0.5)
        minor_grid_alpha = kwargs.get('minor_grid_alpha', 0.0)
        k = kwargs.get('k', 5)
        n_cont = kwargs.get('n_cont', 100)
        n_bins = kwargs.get('n_bins', 10)
        ms = kwargs.get('ms', 10)
        lc = kwargs.get('lc', 'black')
        lw = kwargs.get('lw', 1)
        hc = kwargs.get('hc', 'gray')
        mc = kwargs.get('mc', 'red')
        dist_func = kwargs.get('dist_func', 'pdf')
        fig_ax = kwargs.get('fig_ax', None)

        if fig_ax is None or np.array(fig_ax, dtype='object').any() is None:
            fig = plt.figure(**figkwargs)
            ax = fig.gca()
        else:
            fig = fig_ax[0]
            ax = fig_ax[1]
        ax.minorticks_on()
        ax.grid(True, which="major", alpha=grid_alpha)
        ax.grid(True, which="minor", alpha=minor_grid_alpha)

        if dist_func == 'cdf':
            cumulative = True
        else:
            cumulative = False
        bin_hgts, bin_edges, _ = ax.hist(data, bins=n_bins, density=True, color=hc, edgecolor='white',
                                         cumulative=cumulative)
        _, _, patches = ax.hist(data, bins=n_bins, density=True, color=None, edgecolor=None, cumulative=cumulative,
                                histtype='step', linewidth=0)
        ax.plot(data, 0 * data, '.', color=mc, markersize=ms, zorder=10, clip_on=False)

        export_mat_dict = dict()

        if prob_dist is not None:
            mean = prob_dist.mean()
            std = prob_dist.std()
            x1 = mean - k * std
            x2 = mean + k * std
            x = np.linspace(x1, x2, n_cont)
            if dist_func == 'cdf':
                df_x = prob_dist.cdf(x)
            else:
                df_x = prob_dist.pdf(x)
            ax.plot(x, df_x, color=lc, linewidth=lw)

            xlim = (x1, x2)
            x_bounds = prob_dist.support()
            xlim = (max(xlim[0], x_bounds[0]), min(xlim[1], x_bounds[1]))
            ax.set_xlim(xlim)

            export_mat_dict[f'{dist_func}_plot'] = np.column_stack((x, df_x))
            export_mat_dict['dist_name'] = prob_dist.dist.name

        export_mat_dict['data'] = data
        export_mat_dict['patches'] = patches
        export_mat_dict['bin_hgts'] = bin_hgts
        export_mat_dict['bin_edges'] = bin_edges

        return fig, ax, export_mat_dict

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_rv_x_lims(pd, k=10):
        mean = pd.mean()
        std = pd.std()
        supp = pd.support()
        return max(mean - k * std, supp[0]), min(mean + k * std, supp[-1])

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def shell_exec(exec_file_name, cd_path, exec_commands, **kwargs):
        comp_env = kwargs.get('comp_env', 'local')
        local_bash_path = kwargs.get('local_bash_path', '')
        remove_file = kwargs.get('remove_file', True)

        write_lines = list()
        write_lines.append('#!/bin/bash\n')
        write_lines.append('cd ' + '"' + cd_path + '"\n')
        for command in exec_commands:
            write_lines.append(command + '\n')
        write_lines.append('cd ' + '"' + Utility.get_path(os.getcwd()) + '"\n')

        with open(exec_file_name, 'w') as fid:
            fid.writelines(write_lines)

        if comp_env == 'local' and platform.system() == 'Windows':
            subprocess.call(f"\"{local_bash_path}\" \"{exec_file_name}\"", shell=True)
        else:
            os.system(f'chmod +x "{exec_file_name}"')
            os.system(f'./{exec_file_name}')

        if remove_file:
            os.remove(exec_file_name)

        return

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_curr_time_stamp():
        curr_date_time = datetime.now()
        temp = str(curr_date_time).split()
        curr_date = temp[0]
        curr_time = '-'.join([str(i) for i in [int(float(t)) for t in temp[1].split(':')]])
        return f'date_{curr_date}_time_{curr_time}'

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def untar_file(tar_file_path, **kwargs):

        local_bash_path = kwargs.get('local_bash_path', '')
        comp_env = kwargs.get('comp_env', 'local')
        flags = kwargs.get('flags', ['x', 'z', 'v', 'f'])

        tar_string = f"tar -{''.join(flags)} {os.path.basename(tar_file_path)}"

        exec_commands = [tar_string]
        Utility.shell_exec('UNTAR', os.path.dirname(tar_file_path), exec_commands,
                           comp_env=comp_env, local_bash_path=local_bash_path)

        return

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_rayleigh_damping_params(xi_1, w_1, xi_2, w_2):
        denom = w_1 ** 2 - w_2 ** 2
        alpha = 2. * w_1 * w_2 * (xi_2 * w_1 - xi_1 * w_2) / denom
        beta = (2. * xi_1 * w_1 - 2. * xi_2 * w_2) / denom
        return alpha, beta

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def plot_log_grid(a, axis, **kwargs):
        major_grid_alpha = kwargs.get('major_grid_alpha', 0.5)
        minor_grid_alpha = kwargs.get('minor_grid_alpha', 0.5)
        major_lw = kwargs.get('major_lw', 1)
        minor_lw = kwargs.get('minor_lw', 1)
        major_color = kwargs.get('major_color', 'lightgrey')
        minor_color = kwargs.get('minor_color', 'lightgrey')

        if axis.lower() == 'x':
            axis_lim_power = np.rint(np.log10(np.array(a.get_xlim())))
            other_lim = a.get_ylim()
        elif axis.lower() == 'y':
            axis_lim_power = np.rint(np.log10(np.array(a.get_ylim())))
            other_lim = a.get_xlim()
        else:
            raise ValueError("Invalid axis!")

        major_grid = 10. ** np.arange(axis_lim_power[0], axis_lim_power[1] + 1)
        minor_grid = []
        for i in range(len(major_grid) - 1):
            minor_grid += list(np.linspace(major_grid[i], major_grid[i + 1], 10))
        minor_grid = np.unique(np.array(minor_grid))
        minor_grid = np.array([item for item in minor_grid if item not in major_grid])

        if axis.lower() == 'x':
            for i in range(len(major_grid)):
                a.plot([major_grid[i], major_grid[i]], [other_lim[0], other_lim[1]],
                       lw=major_lw, color=major_color, alpha=major_grid_alpha, zorder=0)
            for i in range(len(minor_grid)):
                a.plot([minor_grid[i], minor_grid[i]], [other_lim[0], other_lim[1]],
                       lw=minor_lw, color=minor_color, alpha=minor_grid_alpha, zorder=0)
        elif axis.lower() == 'y':
            for i in range(len(major_grid)):
                a.plot([other_lim[0], other_lim[1]], [major_grid[i], major_grid[i]],
                       lw=major_lw, color=major_color, alpha=major_grid_alpha, zorder=0)
            for i in range(len(minor_grid)):
                a.plot([other_lim[0], other_lim[1]], [minor_grid[i], minor_grid[i]],
                       lw=minor_lw, color=minor_color, alpha=minor_grid_alpha, zorder=0)

        return a

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_many_dists(dist, params):
        # params must be a 2-d array
        return dist(*params.T)

    # ------------------------------------------------------------------------------------------------------------------

    @staticmethod
    def get_ylim(ax, thresh):
        y_lim_0 = np.inf
        y_lim_1 = -np.inf

        for line in ax.lines:
            y_data = line.get_data()[1]
            if y_data.size == 0:
                continue  # Skip lines with no data
            curr_min = np.min(y_data)
            curr_max = np.max(y_data)
            if curr_min < y_lim_0:
                y_lim_0 = curr_min
            if curr_max > y_lim_1:
                y_lim_1 = curr_max

        # If no valid data was found, return current ylim
        if not np.isfinite(y_lim_0) or not np.isfinite(y_lim_1):
            return ax.get_ylim()

        # Apply thresholds
        ylim = ax.get_ylim()
        if thresh[0] is not None and thresh[1] is None:
            ylim = (y_lim_0 * thresh[0], ylim[1])
        elif thresh[0] is None and thresh[1] is not None:
            ylim = (ylim[0], y_lim_1 * thresh[1])
        elif thresh[0] is not None and thresh[1] is not None:
            ylim = (y_lim_0 * thresh[0], y_lim_1 * thresh[1])

        return ylim
